#!/usr/bin/env python3
"""
АНАЛИЗ СОСТОЯНИЯ ЯЧЕЕК колонки 'Дата приобретения'
Категории: пустые, формулы, даты
"""
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re

def analyze_cell_states_fast(file_path, sheet_name='СК ТПХ_1 пг'):
    """
    БЫСТРЫЙ анализ состояния ячеек: пустые, формулы, готовые даты
    """
    print("🔍 Анализ состояния ячеек колонки 'Дата приобретения'...")
    
    # Сначала определим реальное количество строк с данными
    print("📊 Определение общего количества строк с данными...")
    wb_temp = load_workbook(file_path, data_only=True, read_only=True)
    ws_temp = wb_temp[sheet_name]
    
    total_data_rows = 0
    max_row_to_check = 7000  # Проверяем до 7000 строк
    
    for row_idx, row in enumerate(ws_temp.iter_rows(min_row=12, max_row=max_row_to_check, min_col=1, max_col=3, values_only=True), start=12):
        # Проверяем, есть ли данные в колонках A, B или C
        if any(cell is not None for cell in row):
            total_data_rows = row_idx  # Запоминаем последнюю строку с данными
    
    wb_temp.close()
    
    print(f"📈 Всего строк с данными: {total_data_rows}")
    
    # Загружаем файл только в режиме формул для анализа
    print("📖 Загрузка файла в режиме формул...")
    wb_formulas = load_workbook(file_path, data_only=False, read_only=True)
    ws_formulas = wb_formulas[sheet_name]
    
    # Статистика по категориям
    stats = {
        'empty': 0,
        'formulas': 0,
        'dates': 0,
        'other': 0,
        'formula_examples': [],
        'date_examples': [],
        'other_examples': []
    }
    
    print("📊 Анализ ячеек...")
    
    # Анализируем все строки с данными
    processed = 0
    for row_idx, row in enumerate(ws_formulas.iter_rows(min_row=12, max_row=total_data_rows, min_col=41, max_col=41), start=12):
        if processed % 500 == 0 and processed > 0:
            print(f"   Обработано: {processed}/{total_data_rows - 11} строк...")
        
        if len(row) == 0:
            continue
            
        cell = row[0]  # Колонка AO (41)
        
        # 1. ПУСТЫЕ ячейки
        if cell.value is None:
            stats['empty'] += 1
        
        # 2. ФОРМУЛЫ (начинаются с =)
        elif isinstance(cell.value, str) and cell.value.startswith('='):
            stats['formulas'] += 1
            if len(stats['formula_examples']) < 3:
                stats['formula_examples'].append({
                    'row': row_idx,
                    'formula': cell.value[:100] + "..." if len(cell.value) > 100 else cell.value
                })
        
        # 3. Для остальных типов нужен анализ вычисленных значений
        else:
            # Помечаем как "другое" для now, уточним позже
            stats['other'] += 1
            if len(stats['other_examples']) < 2:
                stats['other_examples'].append({
                    'row': row_idx,
                    'value': str(cell.value)[:50] + "..." if len(str(cell.value)) > 50 else str(cell.value),
                    'type': type(cell.value).__name__
                })
        
        processed += 1
    
    wb_formulas.close()
    
    # Теперь анализируем вычисленные значения ТОЛЬКО для не-формул
    print("📖 Анализ вычисленных значений для не-формульных ячеек...")
    wb_values = load_workbook(file_path, data_only=True, read_only=True)
    ws_values = wb_values[sheet_name]
    
    # Сбрасываем счетчики для переанализа
    actual_dates = 0
    actual_other = 0
    date_examples = []
    other_examples = []
    
    processed_values = 0
    for row_idx, row in enumerate(ws_values.iter_rows(min_row=12, max_row=total_data_rows, min_col=41, max_col=41), start=12):
        if len(row) == 0:
            continue
            
        cell = row[0]
        
        # Пропускаем пустые ячейки
        if cell.value is None:
            processed_values += 1
            continue
        
        # Проверяем соответствующую ячейку в режиме формул
        try:
            # Если это была формула - пропускаем (уже учли)
            formula_cell = ws_formulas.cell(row=row_idx, column=41)
            if (formula_cell.value and 
                isinstance(formula_cell.value, str) and 
                formula_cell.value.startswith('=')):
                processed_values += 1
                continue
        except:
            pass
        
        # 3. ДАТЫ (готовые значения)
        if isinstance(cell.value, (datetime, pd.Timestamp)):
            actual_dates += 1
            if len(date_examples) < 3:
                date_examples.append({
                    'row': row_idx,
                    'date': cell.value
                })
        else:
            actual_other += 1
            if len(other_examples) < 2:
                other_examples.append({
                    'row': row_idx,
                    'value': str(cell.value)[:50] + "..." if len(str(cell.value)) > 50 else str(cell.value),
                    'type': type(cell.value).__name__
                })
        
        processed_values += 1
    
    # Обновляем статистику
    stats['dates'] = actual_dates
    stats['other'] = actual_other
    stats['date_examples'] = date_examples
    stats['other_examples'] = other_examples
    
    wb_values.close()
    
    return stats, total_data_rows - 11  # Возвращаем количество обработанных строк

def main():
    file_path = 'результат.xlsx'
    sheet_name = 'СК ТПХ_1 пг'
    
    print("=" * 70)
    print("🚀 БЫСТРЫЙ АНАЛИЗ СОСТОЯНИЯ ЯЧЕЕК: ПУСТЫЕ • ФОРМУЛЫ • ДАТЫ")
    print("=" * 70)
    
    try:
        # Быстрый анализ состояния ячеек
        stats, total_rows = analyze_cell_states_fast(file_path, sheet_name)
        
        # ВЫВОД РЕЗУЛЬТАТОВ
        print(f"\n🎯 РЕЗУЛЬТАТЫ АНАЛИЗА ({total_rows} строк):")
        print("=" * 50)
        
        # Основная статистика
        categories = [
            ('Пустые ячейки', stats['empty']),
            ('Формулы (не вычисленные)', stats['formulas']),
            ('Готовые даты', stats['dates']),
            ('Другие значения', stats['other'])
        ]
        
        for name, count in categories:
            percentage = (count / total_rows * 100) if total_rows > 0 else 0
            print(f"   {name}: {count} ({percentage:.1f}%)")
        
        # Показать, что еще есть помимо основных категорий
        if stats['formula_examples'] or stats['date_examples'] or stats['other_examples']:
            print(f"\n📊 ПОДРОБНОЕ РАСПРЕДЕЛЕНИЕ:")
            if stats['formula_examples']:
                print(f"   Примеры формул: {len(stats['formula_examples'])} шт.")
            if stats['date_examples']:
                print(f"   Примеры дат: {len(stats['date_examples'])} шт.")
            if stats['other_examples']:
                print(f"   Примеры других значений: {len(stats['other_examples'])} шт.")
        
        # Примеры формул
        if stats['formula_examples']:
            print(f"\n🔧 ПРИМЕРЫ ФОРМУЛ:")
            for ex in stats['formula_examples']:
                print(f"   Строка {ex['row']}: {ex['formula']}")
        
        # Примеры дат
        if stats['date_examples']:
            print(f"\n📅 ПРИМЕРЫ ДАТ:")
            for ex in stats['date_examples']:
                print(f"   Строка {ex['row']}: {ex['date']}")
        
        # Примеры других значений
        if stats['other_examples']:
            print(f"\n❓ ПРИМЕРЫ ДРУГИХ ЗНАЧЕНИЙ:")
            for ex in stats['other_examples']:
                print(f"   Строка {ex['row']}: '{ex['value']}' (тип: {ex['type']})")
        
        # Сводка
        print(f"\n" + "=" * 50)
        print(f"📈 СВОДКА:")
        print(f"   Всего ячеек: {total_rows}")
        print(f"   • Пустых: {stats['empty']}")
        print(f"   • Формул: {stats['formulas']}")
        print(f"   • Дат: {stats['dates']}")
        print(f"   • Других: {stats['other']}")
        
        # Проверка целостности
        total_counted = stats['empty'] + stats['formulas'] + stats['dates'] + stats['other']
        if total_counted == total_rows:
            print(f"✅ ЦЕЛОСТНОСТЬ: Все ячейки учтены")
        else:
            print(f"⚠️  ЦЕЛОСТНОСТЬ: Расхождение ({total_counted} != {total_rows})")
        
        print(f"\n💡 ВЫВОД: В колонке {stats['formulas']} формул и {stats['dates']} готовых дат")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()