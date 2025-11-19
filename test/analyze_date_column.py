#!/usr/bin/env python3
"""
Скрипт для анализа колонки 'Дата приобретения' во вкладке 'СК ТПХ_1 пг'
Определяет позиционирование колонки, количество строк для анализа и 
анализирует типы данных в ячейках по блокам по 500 строк
"""
import re
import math
import datetime


def identify_data_type(value):
    """
    Определяет тип данных значения
    
    Args:
        value: значение ячейки
        
    Returns:
        кортеж (тип_данных, строковое_представление_типа)
    """
    if value is None or (hasattr(value, 'isna') and value.isna()) or str(value).lower() == 'nat':
        return 'empty', 'Пустое значение (NaN/None)'
    
    # Преобразуем в строку для проверки
    str_value = str(value).strip()
    
    # Проверяем на формулу Excel
    if str_value.startswith('='):
        # Выделяем тип формулы, игнорируя параметры
        formula_match = re.match(r'^=([A-Z]+)\s*\(?.*$', str_value)
        if formula_match:
            formula_type = formula_match.group(1)
            return f'formula_{formula_type}', f'Формула: {formula_type}'
        else:
            return 'formula_other', 'Формула: другая'
    
    # Проверяем на ошибку Excel
    if str_value.startswith('#') and str_value.endswith('!'):
        return 'excel_error', f'Ошибка Excel: {str_value}'
    
    # Если уже datetime
    if isinstance(value, datetime.datetime):
        return 'datetime', 'Дата/время (datetime)'
    
    # Если уже timestamp (внутренний формат pandas)
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        try:
            # Пробуем получить атрибуты даты
            year = value.year
            return 'datetime', 'Дата/время (datetime)'
        except:
            pass
    
    # Если число - может быть serial date в Excel
    if isinstance(value, (int, float)):
        if math.isnan(value) or value <= 0:
            return 'numeric_invalid', 'Число: некорректное (NaN/отрицательное)'
        # Excel даты обычно больше 1 (1 января 1900 года)
        if value >= 1:
            try:
                # Попробуем конвертировать Excel serial date в дату
                from datetime import datetime as dt, timedelta
                date_obj = dt(1899, 12, 30) + timedelta(days=value)
                if date_obj.year > 1900 and date_obj.year < 2100:  # Разумный диапазон дат
                    return 'excel_serial_date', 'Excel дата (serial date)'
            except:
                pass
        return 'numeric', 'Число: другое'
    
    # Если строка - проверяем различные форматы дат
    if isinstance(value, str):
        value = value.strip()
        
        # Регулярные выражения для различных форматов дат
        date_patterns = [
            r'^\d{2}\.\d{2}\.\d{4}$',      # DD.MM.YYYY
            r'^\d{4}-\d{2}-\d{2}$',        # YYYY-MM-DD
            r'^\d{2}/\d{2}/\d{4}$',        # DD/MM/YYYY
            r'^\d{2}-\d{2}-\d{4}$',        # DD-MM-YYYY
            r'^\d{4}/\d{2}/\d{2}$',        # YYYY/MM/DD
            r'^\d{1,2}\.\d{1,2}\.\d{4}$',  # D.M.YYYY
            r'^\d{1,2}/\d{1,2}/\d{4}$',    # D/M/YYYY
            r'^\d{1,2}-\d{1,2}-\d{4}$',    # D-M-YYYY
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value):
                # Попробуем распарсить дату
                try:
                    # Импортируем локально, чтобы избежать лишней нагрузки
                    from datetime import datetime as dt
                    dt.strptime(value, get_date_format(pattern))
                    return 'date_string', 'Дата (в строковом формате)'
                except:
                    continue
        
        # Попробуем другие возможные форматы
        possible_formats = [
            '%d.%m.%Y',
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d.%m.%y',
            '%Y-%m-%d %H:%M:%S',  # Включая время
            '%d/%m/%Y %H:%M:%S',
            '%m/%d/%Y',  # Другой формат
            '%m-%d-%Y', # Другой формат
            '%m.%d.%Y',  # Другой формат
        ]
        
        for fmt in possible_formats:
            try:
                from datetime import datetime as dt
                dt.strptime(value, fmt)
                return 'date_string', 'Дата (в строковом формате)'
            except:
                continue
        
        # Если ничего не подошло, попробуем просто распознать дату (ограниченная проверка)
        try:
            # Ограниченная проверка, чтобы не тратить время
            if len(value) >= 8 and any(sep in value for sep in ['.', '-', '/', ' ']):
                # Проверим, содержит ли строка числовые компоненты даты
                import re
                numbers = re.findall(r'\d+', value)
                if len(numbers) >= 3:
                    nums = [int(n) for n in numbers[:3]]
                    # Проверим, выглядит ли как дата (ограниченная проверка)
                    if any(1 <= n <= 31 for n in nums) and any(1900 <= n <= 2100 for n in nums):
                        return 'date_string', 'Дата (в строковом формате)'
        except:
            pass
        
        # Проверяем на числовые строки
        if value.replace('.', '').replace('-', '').isdigit():
            return 'numeric_string', 'Число (в строковом формате)'
        
        # Проверяем на булевы значения
        if value.lower() in ['true', 'false']:
            return 'boolean_string', 'Булево значение (в строковом формате)'
        
        # Остальные строки
        return 'text', 'Текст'
    
    # Если тип не определен, возвращаем как 'other'
    return 'other', f'Другой тип: {type(value).__name__}'


def get_date_format(pattern):
    """Возвращает соответствующий формат для strptime на основе регулярного выражения"""
    if pattern == r'^\d{2}\.\d{2}\.\d{4}$':
        return '%d.%m.%Y'
    elif pattern == r'^\d{4}-\d{2}-\d{2}$':
        return '%Y-%m-%d'
    elif pattern == r'^\d{2}/\d{2}/\d{4}$':
        return '%d/%m/%Y'
    elif pattern == r'^\d{2}-\d{2}-\d{4}$':
        return '%d-%m-%Y'
    elif pattern == r'^\d{4}/\d{2}/\d{2}$':
        return '%Y/%m/%d'
    elif pattern == r'^\d{1,2}\.\d{1,2}\.\d{4}$':
        return '%d.%m.%Y'
    elif pattern == r'^\d{1,2}/\d{1,2}/\d{4}$':
        return '%d/%m/%Y'
    elif pattern == r'^\d{1,2}-\d{1,2}-\d{4}$':
        return '%d-%m-%Y'
    return None


def determine_cell_type(formula_value, value):
    """
    Определяет тип ячейки на основе значения в обоих режимах (формулы и вычисленные значения)
    
    Args:
        formula_value: значение ячейки в режиме формул (data_only=False)
        value: значение ячейки в режиме вычислений (data_only=True)
    
    Returns:
        кортеж (тип_данных, строковое_представление_типа)
    """
    import datetime
    
    # Проверяем на пустое значение
    if value is None or (hasattr(value, 'isna') and value.isna()) or str(value).lower() == 'nat':
        return 'empty', 'Пустое значение (NaN/None)'
    
    # Проверяем на формулу
    if formula_value is not None and isinstance(formula_value, str) and formula_value.startswith('='):
        # Проверим, что это не просто строка, начинающаяся с '=', а именно формула
        if formula_value != value:
            return 'formula', f'Формула: {formula_value}'
    
    # Проверяем на дату
    if isinstance(value, datetime.datetime) or (hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day')):
        try:
            year = value.year
            return 'date', 'Дата/время (datetime)'
        except:
            pass
    
    # Все остальное
    return 'other', f'Другое: {type(value).__name__}'


def analyze_date_column(file_path, column_name='Дата приобретения', sheet_name='СК ТПХ_1 пг', block_size=500):
    """
    Анализирует колонку с датами во вкладке Excel файла по блокам
    
    Args:
        file_path: путь к Excel файлу
        column_name: имя колонки для анализа
        sheet_name: имя листа для анализа
        block_size: размер блока для анализа (по умолчанию 500)
    """
    print(f"Анализ колонки '{column_name}' в файле '{file_path}', лист '{sheet_name}'...")
    
    # ВАЖНО: Загружаем в режиме read_only для экономии памяти
    from openpyxl import load_workbook
    # Открываем файл в двух режимах: с формулами и с вычисленными значениями
    wb_formulas = load_workbook(file_path, data_only=False, read_only=True)
    wb_values = load_workbook(file_path, data_only=True, read_only=True)
    
    ws_formulas = wb_formulas[sheet_name]
    ws_values = wb_values[sheet_name]
    
    # Находим индекс колонки (проверяем строки 10, 11 и 12, как в analyze_dates.py)
    col_index = None
    headers = None
    
    # Проверяем строки 10-12 как в analyze_dates.py
    for row_idx in range(10, 13):
        for row in ws_formulas.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
            for i, header in enumerate(row):
                if header and column_name in str(header):
                    col_index = i
                    headers = row
                    print(f"✅ Найдена колонка '{header}' на позиции: {col_index + 1} (строка заголовка {row_idx}, столбец {chr(65 + col_index)})")
                    break
            if col_index is not None:
                break
        if col_index is not None:
            break
    
    if col_index is None:
        print(f"❌ ОШИБКА: Колонка '{column_name}' не найдена")
        # Показываем доступные заголовки для отладки
        for row_idx in range(10, 13):
            for row in ws_formulas.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
                available_headers = [f"{i+1}:{h}" for i, h in enumerate(row) if h is not None and h != 'None']
                print(f"Заголовки в строке {row_idx}: {available_headers}")
                break
        return None
    
    # Инициализируем словари
    unique_data_types = {}
    type_counts = {}
    
    # Подсчитываем общее количество строк, как в analyze_dates.py
    print("Подсчет общего количества строк...")
    total_rows = 0
    for row in ws_values.iter_rows(min_row=12, values_only=True):
        if len(row) > col_index and row[col_index] is not None:
            total_rows += 1
        elif total_rows > 0:  # Если уже нашли какие-то данные
            # Проверяем, есть ли данные в других колонках в этой строке
            if any(cell is not None for cell in row):
                total_rows += 1  # Все равно считаем строку, если в ней есть какие-то данные
            else:
                # Если строка полностью пустая, предполагаем конец данных
                break
    
    print(f"📊 Всего строк для анализа: {total_rows}")
    
    # ОДИН ПРОХОД по всем строкам с блочным выводом прогресса
    processed_rows = 0
    for row_idx, (formula_row, value_row) in enumerate(zip(
        ws_formulas.iter_rows(min_row=12, values_only=True),
        ws_values.iter_rows(min_row=12, values_only=True)
    ), start=12):
        if processed_rows > 0 and processed_rows % 50 == 0:  # Печатаем каждые 50 строк
            print(f"🔍 Обработано строк: {processed_rows}...")
        
        if len(formula_row) > col_index and len(value_row) > col_index:
            formula_value = formula_row[col_index]
            value = value_row[col_index]
            processed_rows += 1
            
            # Определяем тип данных на основе обоих значений
            data_type_key, data_type_desc = determine_cell_type(formula_value, value)
            
            # Счётчики
            type_counts[data_type_key] = type_counts.get(data_type_key, 0) + 1
            
            # Сохраняем примеры
            if data_type_key not in unique_data_types:
                unique_data_types[data_type_key] = {
                    'description': data_type_desc,
                    'examples': []
                }
            
            if len(unique_data_types[data_type_key]['examples']) < 5:
                unique_data_types[data_type_key]['examples'].append({
                    'formula_value': formula_value,
                    'value': value,
                    'excel_row': row_idx
                })
        
        # Прерываем, если обработали все нужные строки
        if processed_rows >= total_rows:
            break
    
    # Закрываем workbooks
    wb_formulas.close()
    wb_values.close()
    
    # Выводим результаты
    print("\n" + "="*80)
    print("РЕЗУЛЬТАТЫ АНАЛИЗА ТИПОВ ДАННЫХ")
    print("="*80)
    
    print(f"Анализирована колонка: '{column_name}' (столбец {chr(65 + col_index)})")
    print(f"Найдено уникальных типов данных: {len(unique_data_types)}")
    print(f"Всего строк для анализа: {total_rows}")
    print()
    
    # Подсчитываем и выводим количество пустых значений
    empty_count = type_counts.get('empty', 0)
    print(f"Количество пустых значений из колонки {chr(65 + col_index)}: {empty_count} ({empty_count/total_rows*100:.2f}%)")
    print()
    
    for i, (data_type_key, data_type_info) in enumerate(unique_data_types.items(), 1):
        count = type_counts[data_type_key]
        percentage = count / total_rows * 100
        print(f"{i}. {data_type_info['description']}")
        print(f"   Количество: {count} ({percentage:.2f}%)")
        print(f"   Примеры значений:")
        for j, example in enumerate(data_type_info['examples'], 1):
            if example['formula_value'] != example['value']:
                print(f"      {j}. '{example['value']}' (формула: '{example['formula_value']}') (строка Excel {example['excel_row']})")
            else:
                print(f"      {j}. '{example['value']}' (строка Excel {example['excel_row']})")
        print()
    
    print("="*80)
    print("АНАЛИЗ ЗАВЕРШЕН")
    print("="*80)
    
    return unique_data_types


def main():
    file_path = 'результат.xlsx'
    column_name = 'Дата приобретения'
    sheet_name = 'СК ТПХ_1 пг'
    
    try:
        from datetime import timedelta
        results = analyze_date_column(file_path, column_name, sheet_name)
        return results
    except FileNotFoundError:
        print(f"❌ ОШИБКА: Файл {file_path} не найден")
        return None
    except Exception as e:
        print(f"❌ ОШИБКА при анализе файла: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    main()