#!/usr/bin/env python3
"""
ИЗВЛЕЧЕНИЕ НЕПУСТЫХ ЗНАЧЕНИЙ из колонки 'Дата приобретения'
Сохранение в файл по 500-элементных чанков
"""
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os

def extract_non_empty_values(file_path, sheet_name='СК ТПХ_1 пг', chunk_size=500):
    """
    Извлечение всех непустых значений из колонки 'Дата приобретения' по чанкам
    """
    print("🔍 Извлечение непустых значений из колонки 'Дата приобретения'...")
    
    # Сначала определим реальное количество строк с данными
    print("📊 Определение общего количества строк с данными...")
    wb_temp = load_workbook(file_path, data_only=True, read_only=True)
    ws_temp = wb_temp[sheet_name]
    
    total_data_rows = 0
    max_row_to_check = 7000  # Проверяем до 700 строк
    
    for row_idx, row in enumerate(ws_temp.iter_rows(min_row=12, max_row=max_row_to_check, min_col=1, max_col=3, values_only=True), start=12):
        # Проверяем, есть ли данные в колонках A, B или C
        if any(cell is not None for cell in row):
            total_data_rows = row_idx  # Запоминаем последнюю строку с данными
    
    wb_temp.close()
    
    print(f"📈 Всего строк с данными: {total_data_rows}")
    
    # Загружаем файл в режиме формул для анализа
    print("📖 Загрузка файла в режиме формул...")
    wb_formulas = load_workbook(file_path, data_only=False, read_only=True)
    ws_formulas = wb_formulas[sheet_name]
    
    # Собираем непустые значения
    non_empty_values = []
    
    print("📊 Сбор непустых значений...")
    
    # Анализируем все строки с данными
    processed = 0
    for row_idx, row in enumerate(ws_formulas.iter_rows(min_row=12, max_row=total_data_rows, min_col=41, max_col=41), start=12):
        if processed % 500 == 0 and processed > 0:
            print(f"   Обработано: {processed}/{total_data_rows - 11} строк...")
        
        if len(row) == 0:
            continue
            
        cell = row[0]  # Колонка AO (41)
        
        # Пропускаем пустые ячейки
        if cell.value is None:
            continue
        
        # Определяем тип значения
        if isinstance(cell.value, str) and cell.value.startswith('='):
            value_type = 'formula'
            value = cell.value
        else:
            # Для не-формул нужно получить вычисленное значение
            value_type = 'calculated'
            # Загружаем вычисленное значение
            wb_values = load_workbook(file_path, data_only=True, read_only=True)
            ws_values = wb_values[sheet_name]
            calculated_cell = ws_values.cell(row=row_idx, column=41)
            value = calculated_cell.value
            wb_values.close()
        
        non_empty_values.append({
            'row': row_idx,
            'type': value_type,
            'value': value,
            'value_type': type(value).__name__
        })
        
        processed += 1
    
    wb_formulas.close()
    
    print(f"✅ Найдено {len(non_empty_values)} непустых значений")
    
    # Создаем директорию для результатов, если не существует
    output_dir = "results"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Сохраняем непустые значения по чанкам
    chunk_number = 1
    for i in range(0, len(non_empty_values), chunk_size):
        chunk = non_empty_values[i:i + chunk_size]
        
        # Создаем DataFrame для чанка
        df_chunk = pd.DataFrame(chunk)
        
        # Генерируем имя файла
        output_file = os.path.join(output_dir, f"non_empty_values_chunk_{chunk_number}.xlsx")
        
        # Сохраняем чанк в Excel файл
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_chunk.to_excel(writer, sheet_name='non_empty_values', index=False)
            
            # Форматируем заголовки
            worksheet = writer.sheets['non_empty_values']
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Получаем букву колонки
                
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Ограничиваем ширину
                worksheet.column_dimensions[column].width = adjusted_width
        
        print(f"💾 Сохранен чанк {chunk_number}: {output_file} ({len(chunk)} значений)")
        chunk_number += 1
    
    # Также создаем общий файл с полной статистикой
    if non_empty_values:
        df_all = pd.DataFrame(non_empty_values)
        
        # Добавляем статистику
        stats_summary = {
            'total_non_empty': len(non_empty_values),
            'by_type': df_all['type'].value_counts().to_dict(),
            'by_value_type': df_all['value_type'].value_counts().to_dict()
        }
        
        summary_file = os.path.join(output_dir, "non_empty_summary.xlsx")
        with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
            # Основные данные
            df_all.to_excel(writer, sheet_name='all_non_empty', index=False)
            
            # Статистика
            stats_df = pd.DataFrame(list(stats_summary.items()), columns=['metric', 'value'])
            stats_df.to_excel(writer, sheet_name='summary', index=False)
            
            # Статистика по типам в отдельной таблице
            type_stats = []
            for type_name, count in stats_summary['by_type'].items():
                type_stats.append({'type': type_name, 'count': count})
            type_stats_df = pd.DataFrame(type_stats)
            type_stats_df.to_excel(writer, sheet_name='type_statistics', index=False)
            
            # Статистика по типам значений
            value_type_stats = []
            for type_name, count in stats_summary['by_value_type'].items():
                value_type_stats.append({'value_type': type_name, 'count': count})
            value_type_stats_df = pd.DataFrame(value_type_stats)
            value_type_stats_df.to_excel(writer, sheet_name='value_type_statistics', index=False)
        
        print(f"📊 Создан файл со статистикой: {summary_file}")
    
    return non_empty_values

def main():
    file_path = 'результат.xlsx'
    sheet_name = 'СК ТПХ_1 пг'
    
    print("=" * 70)
    print("🚀 ИЗВЛЕЧЕНИЕ НЕПУСТЫХ ЗНАЧЕНИЙ ИЗ КОЛОНКИ 'ДАТА ПРИОБРЕТЕНИЯ'")
    print("=" * 70)
    
    try:
        # Извлечение непустых значений
        non_empty_values = extract_non_empty_values(file_path, sheet_name)
        
        print(f"\n🎯 РЕЗУЛЬТАТЫ:")
        print("=" * 50)
        print(f"   Всего непустых значений: {len(non_empty_values)}")
        
        if non_empty_values:
            # Подсчет по типам
            formulas_count = sum(1 for item in non_empty_values if item['type'] == 'formula')
            calculated_count = sum(1 for item in non_empty_values if item['type'] == 'calculated')
            
            print(f"   Формулы: {formulas_count}")
            print(f"   Вычисленные значения: {calculated_count}")
            
            # Подсчет по типам значений
            print(f"\n📋 Типы значений:")
            value_types = {}
            for item in non_empty_values:
                v_type = item['value_type']
                value_types[v_type] = value_types.get(v_type, 0) + 1
            
            for v_type, count in value_types.items():
                print(f"   {v_type}: {count}")
        
        print(f"\n💾 Файлы с чанками сохранены в директорию 'results/'")
        print(f"📊 Общая статистика сохранена в 'results/non_empty_summary.xlsx'")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()