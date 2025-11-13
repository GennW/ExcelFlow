#!/usr/bin/env python3
"""
Точка входа в приложение ExcelFlow
"""
import argparse
import sys
import time
from pathlib import Path

from processors.data_loader import DataLoader
from processors.parser import DataParser
from processors.matcher import DataMatcher
from utils.logger_config import setup_logger, log_application_start, log_sheet_info, log_processing_progress, log_final_stats


def main():
    parser = argparse.ArgumentParser(description='Приложение для анализа и сопоставления данных Excel')
    parser.add_argument('--input', required=True, help='Путь к входному Excel-файлу')
    parser.add_argument('--output', required=True, help='Путь к выходному Excel-файлу')
    parser.add_argument('--source-sheet', default='ВП 2024-2025 НЧТЗ', help='Имя вкладки-справочника')
    parser.add_argument('--target-sheet', default='СК ТПХ_1 пг', help='Имя вкладки для обработки')
    parser.add_argument('--log-level', default='INFO', help='Уровень логирования (DEBUG, INFO, WARNING, ERROR)')
    parser.add_argument('--report-file', help='Путь к файлу отчёта о несопоставленных записях')
    
    args = parser.parse_args()
    
    # Проверяем, что входной файл существует
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Файл не найден: {args.input}")
        sys.exit(1)
    
    # Настраиваем логгер
    logger = setup_logger(log_level=args.log_level)
    
    # Логируем начало работы
    log_application_start(logger, input_file=args.input, output_file=args.output)
    
    # Запоминаем время начала
    start_time = time.time()
    
    try:
        # Загружаем данные
        loader = DataLoader(logger)
        df_dict = loader.load_excel_file(args.input)
        
        # Логируем информацию о вкладках
        sheet_counts = {sheet: len(df) for sheet, df in df_dict.items()}
        log_sheet_info(logger, list(df_dict.keys()), sheet_counts)
        
        # Парсим и нормализуем данные
        parser = DataParser(logger)
        df_target = df_dict[args.target_sheet]
        df_source = df_dict[args.source_sheet]
        
        logger.info("Начало обработки целевой вкладки...")
        df_target_processed = parser.parse_target_sheet(df_target)
        
        logger.info("Начало обработки справочной вкладки...")
        df_source_processed = parser.parse_source_sheet(df_source)
        
        # Сопоставляем записи
        matcher = DataMatcher(logger)
        logger.info("Начало сопоставления записей...")
        
        df_result, stats = matcher.match_records(df_target_processed, df_source_processed)
        
        # Обновляем словарь с DataFrame
        df_dict[args.target_sheet] = df_result
        
        # Сохраняем результат
        logger.info(f"Сохранение результата в файл: {args.output}")
        
        # Используем openpyxl для сохранения с сохранением форматирования
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Загружаем исходный файл для сохранения форматирования
        wb = load_workbook(args.input)
        
        # Обновляем целевую вкладку
        ws = wb[args.target_sheet]
        
        # Определяем диапазон для вставки новых данных
        start_row = 2  # Пропускаем заголовки
        start_col = len(df_target.columns) + 1  # Начинаем с первой пустой колонки
        
        # Добавляем заголовки новых столбцов
        ws.cell(row=1, column=start_col, value="Дата приобретения")
        ws.cell(row=1, column=start_col + 1, value="Квартал приобретения")
        ws.cell(row=1, column=start_col + 2, value="Рассчитанная себестоимость")
        
        # Заполняем данные
        for row_idx, row_data in df_result.iterrows():
            # Форматируем дату приобретения
            acquisition_date = row_data['Дата приобретения']
            if acquisition_date:
                date_str = acquisition_date.strftime('%d.%m.%Y')
            else:
                date_str = '*Дата не найдена*'
                
            ws.cell(row=row_idx + 2, column=start_col, value=date_str)
            ws.cell(row=row_idx + 2, column=start_col + 1, value=row_data['Квартал приобретения'])
            ws.cell(row=row_idx + 2, column=start_col + 2, value=row_data['Рассчитанная себестоимость'])
        
        # Сохраняем в новый файл
        try:
            wb.save(args.output)
            logger.info(f"Файл успешно сохранен: {args.output}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении файла: {str(e)}")
            import traceback
            logger.error(f"Детали ошибки сохранения: {traceback.format_exc()}")
            sys.exit(6)
        
        # Логируем итоговую статистику
        total_records = len(df_target)
        matched_records = total_records - stats['manual_checks'] - stats['missing_data']
        
        processing_time = time.time() - start_time
        
        # Добавляем мониторинг производительности
        import psutil
        import os
        memory_usage = psutil.Process().memory_info().rss / 1024 / 1024  # в МБ
        logger.info(f"Использование памяти в конце обработки: {memory_usage:.2f} MB")
        
        log_final_stats(
            logger,
            total_records=total_records,
            matched_records=matched_records,
            manual_check_records=stats['manual_checks'],
            missing_data_records=stats['missing_data'],
            exact_matches=stats['exact_matches'],  # будет 0, так как точные совпадения не отдельно считаются
            level1_matches=stats['level1_matches'],
            level2_matches=stats['level2_matches'],
            fuzzy_matches=stats['fuzzy_matches'],
            processing_time=processing_time
        )
        
        logger.info("Обработка завершена успешно!")
        
    except Exception as e:
        logger.error(f"Ошибка при обработке: {str(e)}")
        import traceback
        logger.error(f"Детали ошибки: {traceback.format_exc()}")
        sys.exit(5)


if __name__ == "__main__":
    main()