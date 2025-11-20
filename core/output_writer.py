"""Запись результатов в Excel (оптимизировано для производительности)"""
import pandas as pd
from openpyxl import load_workbook
from utils.logger import get_logger

logger = get_logger(__name__)

def write_results(input_file: str, output_file: str, df_result: pd.DataFrame, target_sheet: str) -> None:
    """
    Оптимизированная запись результатов в Excel.
    
    Заголовки записываются на строку 11 (где находятся заголовки существующих столбцов).
    Данные записываются начиная со строки 12 (где начинаются данные существующих столбцов).
    
    Оптимизация: используем списки вместо многократного обращения к DataFrame,
    массовая запись через openpyxl без создания промежуточных объектов.
    
    Args:
        input_file: Путь к входному Excel-файлу
        output_file: Путь к выходному Excel-файлу
        df_result: DataFrame с результатами обработки
        target_sheet: Название целевого листа
    """
    try:
        wb = load_workbook(input_file)
        ws = wb[target_sheet]
        logger.info(f"Открыт файл для записи: {input_file}")

        # Определяем первый свободный столбец после существующих данных
        start_col = ws.max_column + 1
        
        # Заголовки новых столбцов
        headers = [
            '**Дата приобретения**',
            '**Квартал приобретения**',
            '**Стоимость закупки НЧТЗ 1 ед**',
            '**Прямая СС НЧТЗ 1 ед**',
            '**НР НЧТЗ 1 ед**'
        ]

        # Записываем заголовки в строку 11 (где находятся заголовки существующих столбцов)
        for i, header in enumerate(headers):
            ws.cell(row=11, column=start_col + i, value=header)

        logger.info(f"Добавлены заголовки в строку 11, столбцы {start_col}-{start_col + 4}")

        # Оптимизация: извлекаем данные в списки один раз, а не при каждой итерации
        ao_values = df_result['AO_**Дата_приобретения**'].tolist()
        ap_values = df_result['AP_**Квартал_приобретения**'].tolist()
        aq_values = df_result['AQ_**Стоимость_закупки**'].tolist()
        ar_values = df_result['AR_**Прямая_СС**'].tolist()
        as_values = df_result['AS_**НР**'].tolist()

        # Массовая запись данных начиная со строки 12 (где начинаются данные существующих столбцов)
        for idx in range(len(ao_values)):
            row_num = 12 + idx  # Данные начинаются со строки 12
            
            ws.cell(row=row_num, column=start_col,     value=ao_values[idx])
            ws.cell(row=row_num, column=start_col + 1, value=ap_values[idx])
            ws.cell(row=row_num, column=start_col + 2, value=aq_values[idx])
            ws.cell(row=row_num, column=start_col + 3, value=ar_values[idx])
            ws.cell(row=row_num, column=start_col + 4, value=as_values[idx])

        logger.info(f"Записано {len(ao_values)} строк данных, начиная со строки 12")
        wb.save(output_file)
        logger.info(f"Файл успешно сохранён: {output_file}")
        
    except Exception as e:
        logger.error(f"Ошибка при записи: {e}")
        raise
